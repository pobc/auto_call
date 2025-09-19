import openpyxl

# 加载 Excel 文件
input_file = r"C:\Users\jiang\Documents\海居租房\茉莉公馆4栋\副本茉莉公馆--编辑.xlsx"  # 替换为你的 Excel 文件路径
output_file = r"C:\Users\jiang\Documents\海居租房\茉莉公馆4栋\output.xlsx"  # 输出文件路径
wb = openpyxl.load_workbook(input_file)
ws = wb.active  # 假设数据在第一个工作表

# 遍历工作表的所有行（从第 1 行开始，假设没有标题行）
for row in range(1, ws.max_row + 1):
    # 获取 E 列（第 5 列）、F 列（第 6 列）、G 列（第 7 列）的单元格值
    e_cell = ws.cell(row=row, column=5).value
    f_cell = ws.cell(row=row, column=6).value
    g_cell = ws.cell(row=row, column=7).value

    # 第一步：如果 F 列有数据且 E 列为空，将 F 列数据移动到 E 列
    # if f_cell is not None and e_cell is None:
    #     ws.cell(row=row, column=5).value = f_cell  # 复制 F 列到 E 列
    #     ws.cell(row=row, column=6).value = None    # 清空 F 列

    # 第二步：检查 F 列是否为空，如果为空且 G 列有数据，将 G 列数据移动到 F 列
    f_cell_updated = ws.cell(row=row, column=6).value  # 重新获取 F 列值
    if f_cell_updated is None and g_cell is not None:
        ws.cell(row=row, column=6).value = g_cell  # 复制 G 列到 F 列
        ws.cell(row=row, column=7).value = None    # 清空 G 列

# 保存修改后的 Excel 文件
wb.save(output_file)
print(f"处理完成，结果已保存到 {output_file}")